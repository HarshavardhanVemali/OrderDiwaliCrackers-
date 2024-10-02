from django.shortcuts import render,redirect
from .models import Items,Order,OrderItem,FailedLoginAttempts
from django.http import JsonResponse, HttpResponse 
import json
from django.db.models import Sum
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from django.urls import reverse
from django.contrib.auth import logout
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import user_passes_test
import uuid
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook
from io import BytesIO 

def admin_required(view_func):
    return user_passes_test(
        lambda u: u.is_superuser or u.has_perm('queryraiserapp.is_admin'), 
        login_url='/adminlogin/' 
    )(view_func)

MAX_FAILED_ATTEMPTS = 3
def set_device_cookie(response, device_id):
    response.set_cookie('device_id', device_id, max_age=365*24*60*60)

def get_device_id(request):
    return request.COOKIES.get('device_id', str(uuid.uuid4()))

def is_device_blocked(device_id):
    try:
        failed_attempt = FailedLoginAttempts.objects.get(device_id=device_id)
        if failed_attempt.is_active:
            return True
    except FailedLoginAttempts.DoesNotExist:
        return False
    return False

def adminlogin(request):
    if request.method == 'POST':
        device_id = get_device_id(request)
        if is_device_blocked(device_id):
            return render(request, 'adminlogin.html', {'blocked': True, 'error_message': 'Your device is permanently blocked due to multiple failed login attempts.'})

        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            if user.is_superuser or (user.is_staff and not user.is_active):
                FailedLoginAttempts.objects.filter(device_id=device_id).update(attempts=0, is_active=False)
                response = redirect('admindashboard')
                set_device_cookie(response, device_id)
                login(request, user)
                return response
            else:
                return render(request, 'adminlogin.html', {'error_message': 'You do not have permission to access the admin page.'})
        else:
            failed_attempt, created = FailedLoginAttempts.objects.get_or_create(device_id=device_id)
            if not created:
                failed_attempt.attempts += 1
                failed_attempt.save()
                if failed_attempt.attempts >= MAX_FAILED_ATTEMPTS:
                    failed_attempt.is_active = True
                    failed_attempt.save()
                    return render(request, 'adminlogin.html', {'blocked': True, 'error_message': 'Your device is permanently blocked due to multiple failed login attempts.'})
            else:
                failed_attempt.attempts = 1
                failed_attempt.save()
            return render(request, 'adminlogin.html', {'error_message': 'Invalid username or password'})
    else:
        response = render(request, 'adminlogin.html')
        if 'device_id' not in request.COOKIES:
            device_id = get_device_id(request)
            set_device_cookie(response, device_id)
        return response

def index(request):
    return render(request,'index.html')

def adminloginpage(request):
    return render(request,'adminlogin.html')


@admin_required
def admindashboard(request):
    return render(request,'admindashboard.html')

def admingetitems(request):
    if request.method == 'POST':
        items = Items.objects.all().values('item_number', 'item_name', 'price')
        items_list = list(items)
        return JsonResponse(items_list, safe=False)
    else:
        return JsonResponse({'success':False,'error':'Invalid request method.'}) 

@admin_required
def add_item(request):
    if request.method == 'POST':
        item_name = request.POST.get('itemname')
        price = request.POST.get('price')

        if item_name and price:
            try:
                price = float(price) 
                new_item = Items(item_name=item_name, price=price)
                new_item.save()
                return JsonResponse({'success': True})
            except ValueError:
                return JsonResponse({'success': False, 'error': 'Invalid price format'})
        else:
            return JsonResponse({'success': False, 'error': 'Missing item name or price'})
    else:
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
@admin_required
def delete_item(request):
    if request.method == 'POST':
        data=json.loads(request.body)
        item_number = data.get('item_number')
        
        if item_number:
            try:
                item = Items.objects.get(item_number=item_number)
                item.delete()
                return JsonResponse({'success': True})
            except Items.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Item not found'})
        else:
            return JsonResponse({'success': False, 'error': 'Missing item number'})
    else:
        return JsonResponse({'success': False, 'error': 'Invalid request method'})


def place_order(request):
    if request.method == 'POST':
        customer_name = request.POST.get('name')
        phone_number = request.POST.get('phone')
        email = request.POST.get('email')
        address = request.POST.get('address')

        item_numbers = request.POST.getlist('item_number')

        quantities = []
        for item_number in item_numbers:
            quantity = request.POST.get(f'quantity_{item_number}')
            if quantity:
                quantities.append(int(quantity))

        if customer_name and phone_number and address and item_numbers and quantities:
            if not any(quantity > 0 for quantity in quantities):
                return JsonResponse({'success': False, 'error': 'At least one item quantity must be greater than 0.'})
            try:
                total_amount = 0
                order = Order.objects.create(
                    customer_name=customer_name,
                    phone_number=phone_number,
                    email=email,
                    address=address,
                    total_amount=total_amount  
                )
                for item_number, quantity in zip(item_numbers, quantities):
                    try:
                        item = Items.objects.get(item_number=item_number)
                        quantity = int(quantity)
                        if quantity > 0:
                            item_total = item.price * quantity
                            total_amount += item_total
                            OrderItem.objects.create(
                                order=order,
                                item=item,
                                quantity=quantity,
                                item_price=item.price
                            )
                    except Items.DoesNotExist:
                        return JsonResponse({'success': False, 'error': f'Item with ID {item_number} not found.'})
                    except ValueError:
                        return JsonResponse({'success': False, 'error': f'Invalid quantity for item {item_number}.'})
                order.total_amount = total_amount
                order.save()
                invoice_url = reverse('generate_invoice_pdf', kwargs={'order_id': order.order_id})
                return JsonResponse({'success': True, 'message': 'Order placed successfully!', 'invoice_url': invoice_url})

            except Exception as e:
                return JsonResponse({'success': False, 'error': f'Error placing order: {str(e)}'})

        else:
            return JsonResponse({'success': False, 'error': 'Please fill all required fields and select at least one item.'})
    else:
        return JsonResponse({'success': False, 'error': 'Invalid request method.'})

@admin_required
def adminorderpage(request):
    return render(request,'adminorderpage.html')

@admin_required
def getorders(request):
    if request.method == 'POST':
        orders = Order.objects.all().values('order_id', 'customer_name', 'phone_number', 'address','status')
        orders_list = list(orders) 
        return JsonResponse(orders_list, safe=False)
    else:
        return JsonResponse({'success': False, 'error': 'Invalid request method.'})

@admin_required
def getorderitems(request, order_id):
    if request.method == 'POST':
        try:
            order = Order.objects.get(order_id=order_id) 

            items_data = []
            for order_item in order.items.all():
                item_data = {
                    'item_name': order_item.item.item_name,
                    'quantity': order_item.quantity,
                    'item_price': float(order_item.item_price), 
                    'item_total': float(order_item.item_price * order_item.quantity),  
                }
                items_data.append(item_data)

            response_data = {
                'items': items_data,
                'total_amount': float(order.total_amount)  
            }
            return JsonResponse(response_data, safe=False)
        except Order.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Order not found.'})

    else:
        return JsonResponse({'success': False, 'error': 'Invalid request method.'})

@admin_required
def orderdelevired(request,order_id):
    if request.method=='POST':
        try:
            order = Order.objects.get(order_id=order_id) 
            order.status='delivered'
            order.save()
            return JsonResponse({'success':True})
        except Order.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Order not found.'})
    else:
        return JsonResponse({'success': False, 'error': 'Invalid request method.'})

@admin_required
def deleteorder(request,order_id):
    if request.method=='POST':
        try:
            order = Order.objects.get(order_id=order_id)     
            order.delete()
            return JsonResponse({'success':True})
        except Order.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Order not found.'})
    else:
        return JsonResponse({'success': False, 'error': 'Invalid request method.'})
    
@admin_required
def download_report(request):
    if request.method == 'POST':
        try:
            item_counts = OrderItem.objects.values('item__item_number', 'item__item_name').annotate(
                total_quantity=Sum('quantity')
            ).order_by('-total_quantity')
            wb = Workbook()
            ws = wb.active
            ws.title = "Item Counts Report" 
            headers = [
                'Item Number',
                'Item Name',
                'Total Quantity'
            ]
            for col_num, header in enumerate(headers, 1):  
                ws.cell(row=1, column=col_num, value=header)
                ws.column_dimensions[chr(ord('A') + col_num - 1)].width = 20 
            for row_num, item in enumerate(item_counts, 2):  
                ws.cell(row=row_num, column=1, value=item['item__item_number'])
                ws.cell(row=row_num, column=2, value=item['item__item_name'])
                ws.cell(row=row_num, column=3, value=item['total_quantity'])

            output = BytesIO()
            wb.save(output)  
            output.seek(0) 

            response = HttpResponse(
                output.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response['Content-Disposition'] = 'attachment; filename=Item_Counts_Report.xlsx'

            return response

        except Exception as e: 
            return JsonResponse({'error': f'An error occurred: {str(e)}'}, status=500)
    else:
        return JsonResponse({'error': 'Method not allowed. Use POST.'}, status=405)
    

@admin_required
def adminitemcountpage(request):
    return render(request,'adminitemcount.html')

@admin_required
def get_item_counts(request):
    if request.method == 'POST':
        item_counts = OrderItem.objects.values('item__item_number', 'item__item_name').annotate(
            total_quantity=Sum('quantity')
        ).order_by('-total_quantity') 
        return JsonResponse(list(item_counts), safe=False)
    else:
        return JsonResponse({'success': False, 'error': 'Invalid request method.'})

def generate_invoice_pdf(request, order_id):
    try:
        order = Order.objects.get(order_id=order_id)
    except Order.DoesNotExist:
        return HttpResponse("Order not found.", status=404)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="order_{order_id}_invoice.pdf"'
    p = canvas.Canvas(response, pagesize=letter)
    p.drawString(1 * inch, 10.5 * inch, "Diwali Crackers")
    p.drawString(1 * inch, 10 * inch, "Invoice")
    p.line(1 * inch, 9.8 * inch, 7 * inch, 9.8 * inch)
    p.drawString(1 * inch, 9.5 * inch, f"Order ID: {order.order_id}")
    p.drawString(1 * inch, 9 * inch, f"Recipient Name: {order.customer_name}")
    p.drawString(1 * inch, 8.5 * inch, f"Phone Number: {order.phone_number}")
    p.drawString(1 * inch, 8 * inch, f"Address: {order.address}")
    def add_page_header():  
        p.drawString(1 * inch, 7.5 * inch, "Item Name") 
        p.drawString(4.5 * inch, 7.5 * inch, "Quantity")  
        p.drawString(5.5 * inch, 7.5 * inch, "Price")     
        p.drawString(6.5 * inch, 7.5 * inch, "Total")    
        p.line(1 * inch, 7.3 * inch, 7 * inch, 7.3 * inch)

    y = 7 * inch

    add_page_header()
    for item in order.items.all():
        if y < 1 * inch: 
            p.showPage()
            y = 7 * inch 
            add_page_header()

        p.drawString(1 * inch, y, item.item.item_name)  
        p.drawString(4.5 * inch, y, str(item.quantity))  
        p.drawString(5.5 * inch, y, f"Rs.{item.item_price}")  
        p.drawString(6.5 * inch, y, f"Rs.{item.item_price * item.quantity}") 
        y -= 0.3 * inch

    if y < 1 * inch:  
        p.showPage()
        y = 7 * inch
        add_page_header()

    p.line(1 * inch, y - 0.2 * inch, 7 * inch, y - 0.2 * inch)
    p.drawString(5.5 * inch, y - 0.5 * inch, "Grand Total:")
    p.drawString(6.5 * inch, y - 0.5 * inch, f"Rs.{order.total_amount}")

    p.showPage()
    p.save()

    return response

@admin_required
def admin_logout_view(request):
    logout(request)
    return redirect('index')
